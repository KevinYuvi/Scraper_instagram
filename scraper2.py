import os
import json
import asyncio
import re
from datetime import datetime
from typing import List, Dict

from dotenv import load_dotenv
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()


class InstagramStatsScraper:

    def __init__(self):
        self.username = os.getenv("INSTAGRAM_USERNAME")
        self.password = os.getenv("INSTAGRAM_PASSWORD")
        self.profile_url = os.getenv("INSTAGRAM_PROFILE_URL")
        self.max_posts = int(os.getenv("MAX_POSTS", 10))
        self.headless = os.getenv("HEADLESS_MODE", "false").lower() == "true"
        self.exclude_pinned = os.getenv("EXCLUDE_PINNED", "false").lower() == "true"

        self.browser = None
        self.context = None
        self.page = None
        self.playwright = None

        self.session_file = "instagram_session.json"

    # ======================================================
    # BROWSER
    # ======================================================

    async def init(self):
        self.playwright = await async_playwright().start()

        self.browser = await self.playwright.chromium.launch(
            headless=self.headless
        )

    async def create_context(self, storage_state=None):
        self.context = await self.browser.new_context(
            storage_state=storage_state,
            viewport={"width": 1400, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/135.0.0.0 Safari/537.36"
            )
        )

        self.page = await self.context.new_page()

    # ======================================================
    # SESSION / COOKIES
    # ======================================================

    async def load_session(self):
        if not os.path.exists(self.session_file):
            return False

        try:
            with open(self.session_file, "r", encoding="utf-8") as f:
                storage = json.load(f)

            await self.create_context(storage_state=storage)

            await self.page.goto("https://www.instagram.com/")
            await asyncio.sleep(3)

            if "accounts/login" not in self.page.url:
                print("✅ Cookies cargadas")
                return True

            return False

        except Exception:
            return False

    async def save_session(self):
        storage = await self.context.storage_state()

        with open(self.session_file, "w", encoding="utf-8") as f:
            json.dump(storage, f, indent=2)

        print("🍪 Cookies guardadas")

    async def login(self):
        await self.create_context()

        await self.page.goto("https://www.instagram.com/accounts/login/")
        await asyncio.sleep(4)

        await self.page.fill('input[name="username"]', self.username)
        await self.page.fill('input[name="password"]', self.password)
        await self.page.click('button[type="submit"]')

        await asyncio.sleep(7)
        await self.save_session()

        print("✅ Login realizado")

    async def setup_session(self):
        await self.init()

        loaded = await self.load_session()
        if loaded:
            return

        print("🔐 Iniciando nueva sesión...")
        await self.login()

    # ======================================================
    # HELPERS
    # ======================================================

    def parse_count(self, text):
        text = str(text).upper().strip().replace(" ", "")

        try:
            if "MILL" in text:
                num = re.sub(r"[^0-9,\.]", "", text).replace(".", "").replace(",", ".")
                return int(float(num) * 1_000_000)

            if "MIL" in text:
                num = re.sub(r"[^0-9,\.]", "", text).replace(".", "").replace(",", ".")
                return int(float(num) * 1_000)

            if text.endswith("M"):
                return int(float(text[:-1].replace(",", ".")) * 1_000_000)

            if text.endswith("K"):
                return int(float(text[:-1].replace(",", ".")) * 1_000)

            return int(text.replace(".", "").replace(",", ""))

        except Exception:
            return 0

    def extract_hashtags(self, text):
        if not text:
            return []

        tags = re.findall(r'#[\wáéíóúÁÉÍÓÚñÑ_]+', text)
        return sorted(list(set(tags)))

    # ======================================================
    # GET POSTS
    # ======================================================

    async def get_posts(self):
        print(f"\n📱 Abriendo perfil: {self.profile_url}")

        await self.page.goto(self.profile_url, wait_until="domcontentloaded")
        await asyncio.sleep(5)

        for _ in range(4):
            await self.page.mouse.wheel(0, 1800)
            await asyncio.sleep(1.2)

        items = await self.page.evaluate("""
        () => {
            const anchors = Array.from(document.querySelectorAll('a[href*="/p/"], a[href*="/reel/"]'));

            return anchors
                .filter(a => a.querySelector('img'))
                .map((a, idx) => {
                    const rect = a.getBoundingClientRect();

                    const text = (a.innerText || '').toLowerCase();
                    const aria = (a.getAttribute('aria-label') || '').toLowerCase();
                    const html = (a.outerHTML || '').toLowerCase();

                    const pinned =
                        text.includes('anclada') ||
                        text.includes('pinned') ||
                        aria.includes('anclada') ||
                        aria.includes('pinned') ||
                        html.includes('anclada') ||
                        html.includes('pinned');

                    return {
                        href: a.href || a.getAttribute('href'),
                        top: Math.round(rect.top + window.scrollY),
                        left: Math.round(rect.left + window.scrollX),
                        idx,
                        pinned
                    };
                })
                .filter(x => x.href)
                .sort((a, b) => {
                    if (Math.abs(a.top - b.top) > 30) return a.top - b.top;
                    if (a.left !== b.left) return a.left - b.left;
                    return a.idx - b.idx;
                });
        }
        """)

        final = []
        seen = set()

        for item in items:
            url = item["href"]
            if not url:
                continue

            url = url.split("?")[0].strip()

            if url.startswith("/"):
                url = "https://www.instagram.com" + url

            if "/p/" not in url and "/reel/" not in url:
                continue

            if url in seen:
                continue

            seen.add(url)

            final.append({
                "url": url,
                "pinned": item.get("pinned", False)
            })

        if self.exclude_pinned:
            final = [x for x in final if not x["pinned"]]

        print(f"✅ Publicaciones detectadas: {len(final)}")
        for i, item in enumerate(final[:min(8, len(final))], 1):
            pin = "📌" if item["pinned"] else "  "
            print(f"{pin} {i}. {item['url']}")

        return final[:self.max_posts]

    # ======================================================
    # SCRAPE SINGLE POST
    # ======================================================

    async def scrape_post(self, url, idx):
        await self.page.goto(url)
        await asyncio.sleep(4)

        result = {
            "index": idx + 1,
            "tipo": "reel" if "/reel/" in url else "post",
            "anclada": "",
            "usuario": "",
            "fecha": "",
            "likes": 0,
            "comentarios": 0,
            "hashtags": [],
            "url": url,
            "caption": ""
        }

        # usuario
        try:
            links = await self.page.query_selector_all("header a[href^='/']")

            for el in links:
                txt = (await el.inner_text()).strip()
                href = await el.get_attribute("href")

                if txt and href and "/p/" not in href and "/reel/" not in href:
                    result["usuario"] = txt
                    break
        except Exception:
            pass

        # fecha
        try:
            time_el = await self.page.query_selector("time")
            if time_el:
                result["fecha"] = await time_el.get_attribute("datetime")
        except Exception:
            pass

        # texto visible
        text = ""

        for selector in ["article", "main", "body"]:
            try:
                t = await self.page.inner_text(selector)
                if t:
                    text += "\n" + t
            except Exception:
                pass

        # likes
        patterns = [
            r'(\d+[.,]?\d*\s*mill\.)',
            r'(\d+[.,]?\d*\s*mil)',
            r'(\d+[.,]?\d*\s*M)\b',
            r'(\d+[.,]?\d*\s*K)\b'
        ]

        for p in patterns:
            m = re.search(p, text, re.I)
            if m:
                result["likes"] = self.parse_count(m.group(1))
                break

        # comentarios totales
        result["comentarios"] = 0

        # 1) Prioridad: JSON-LD del post actual
        try:
            ld_scripts = await self.page.query_selector_all('script[type="application/ld+json"]')
            for sc in ld_scripts:
                raw = await sc.inner_text()
                if not raw:
                    continue

                try:
                    parsed = json.loads(raw)
                except Exception:
                    continue

                items = parsed if isinstance(parsed, list) else [parsed]

                for item in items:
                    if not isinstance(item, dict):
                        continue

                    if item.get("commentCount") is not None:
                        result["comentarios"] = int(item["commentCount"])
                        break

                    stats = item.get("interactionStatistic")
                    if isinstance(stats, dict):
                        if stats.get("@type") == "CommentAction" and stats.get("userInteractionCount") is not None:
                            result["comentarios"] = int(stats["userInteractionCount"])
                            break
                    elif isinstance(stats, list):
                        for stat in stats:
                            if not isinstance(stat, dict):
                                continue
                            if stat.get("@type") == "CommentAction" and stat.get("userInteractionCount") is not None:
                                result["comentarios"] = int(stat["userInteractionCount"])
                                break

                if result["comentarios"] > 0:
                    break
        except Exception:
            pass

        # 2) Respaldo: scripts filtrados por shortcode del post actual
        if result["comentarios"] == 0:
            try:
                shortcode = url.rstrip("/").split("/")[-1]

                scripts = await self.page.query_selector_all("script")
                for sc in scripts:
                    try:
                        content = await sc.inner_text()
                        if not content or shortcode not in content:
                            continue

                        patterns = [
                            r'"comment_count"\s*:\s*(\d+)',
                            r'"edge_media_to_parent_comment"\s*:\s*\{.*?"count"\s*:\s*(\d+)',
                            r'"edge_media_to_comment"\s*:\s*\{.*?"count"\s*:\s*(\d+)'
                        ]

                        for pattern in patterns:
                            m = re.search(pattern, content, re.DOTALL)
                            if m:
                                result["comentarios"] = int(m.group(1))
                                break

                        if result["comentarios"] > 0:
                            break
                    except Exception:
                        continue
            except Exception:
                pass

        # caption
        lines = [x.strip() for x in text.split("\n") if x.strip()]
        candidates = []

        for line in lines:
            if len(line) < 15:
                continue

            if re.fullmatch(r'[\d.,]+\s*(mil|mill\.|M|K)?', line, re.I):
                continue

            if result["usuario"] and line.lower() == result["usuario"].lower():
                continue

            candidates.append(line)

        if candidates:
            result["caption"] = max(candidates, key=len)

        # hashtags
        result["hashtags"] = self.extract_hashtags(text + " " + result["caption"])

        print(
            f"✅ {idx+1} | Likes: {result['likes']} | "
            f"Comentarios: {result['comentarios']} | "
            f"Hashtags: {len(result['hashtags'])}"
        )

        return result

    # ======================================================
    # EXCEL
    # ======================================================

    def save_excel(self, rows):
        os.makedirs("results", exist_ok=True)

        file = f"results/instagram_stats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Posts"

        ws.append([
            "Index",
            "Tipo",
            "Anclada",
            "Usuario",
            "Fecha",
            "Likes",
            "Comentarios",
            "Hashtags",
            "URL",
            "Caption"
        ])

        fill = PatternFill("solid", fgColor="1F4E78")
        font = Font(color="FFFFFF", bold=True)

        for c in ws[1]:
            c.fill = fill
            c.font = font
            c.alignment = Alignment(horizontal="center", vertical="center")

        for r in rows:
            ws.append([
                r["index"],
                r["tipo"],
                r.get("anclada", "no"),
                r["usuario"],
                r["fecha"],
                r["likes"],
                r["comentarios"],
                ", ".join(r["hashtags"]),
                r["url"],
                r["caption"]
            ])

        widths = {
            "A": 8,
            "B": 10,
            "C": 10,
            "D": 20,
            "E": 25,
            "F": 14,
            "G": 14,
            "H": 40,
            "I": 55,
            "J": 100
        }

        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        for row in ws.iter_rows(min_row=2):
            row[7].alignment = Alignment(wrap_text=True, vertical="top")
            row[9].alignment = Alignment(wrap_text=True, vertical="top")

        ws.freeze_panes = "A2"

        wb.save(file)
        return file

    # ======================================================
    # MAIN
    # ======================================================

    async def run(self):
        await self.setup_session()

        posts = await self.get_posts()
        rows = []

        for i, item in enumerate(posts):
            try:
                row = await self.scrape_post(item["url"], i)
                row["anclada"] = "sí" if item.get("pinned") else "no"
                rows.append(row)
                await asyncio.sleep(2)
            except Exception as e:
                print("⚠️ Error:", e)

        file = self.save_excel(rows)

        print("\n📁 Archivo generado:", file)

        await self.browser.close()
        await self.playwright.stop()


async def main():
    scraper = InstagramStatsScraper()
    await scraper.run()


if __name__ == "__main__":
    asyncio.run(main())