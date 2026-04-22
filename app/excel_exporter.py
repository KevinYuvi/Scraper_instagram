from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class ExcelExporter:
    def save(self, rows, output_dir) -> str:
        output_dir.mkdir(parents=True, exist_ok=True)
        file_path = output_dir / f"instagram_stats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Posts"

        ws.append([
            "Index", "Tipo", "Fecha",
            "Likes", "Comentarios", "Hashtags", "URL", "Caption"
        ])

        fill = PatternFill("solid", fgColor="1F4E78")
        font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in rows:
            ws.append([
                row.index,
                row.tipo,
                row.fecha,
                row.likes,
                row.comentarios,
                ", ".join(row.hashtags),
                row.url,
                row.caption,
            ])

        wb.save(file_path)
        return str(file_path)